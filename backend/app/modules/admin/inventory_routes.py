from __future__ import annotations

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Path, Response
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, delete
from sqlalchemy.ext.asyncio import AsyncSession
from io import BytesIO, StringIO
from openpyxl import Workbook, load_workbook

from app.core.deps import get_admin_actor, require_admin_key
from app.core.db import get_db
from app.models.category import Category
from app.models.item import Item
from app.models.item_unit import ItemUnit
from app.models.industry import Industry, IndustryArticle
from app.modules.inventory.routes import _normalize_sku, CSV_COLUMNS
from app.modules.inventory.schemas import (
    CategoryCreate,
    CategoryOut,
    CategoryUpdate,
    ItemCreate,
    ItemOut,
    ItemUpdate,
    ItemsPage,
    ItemUnitOut,
    IndustryCreate,
    IndustryOut,
    IndustryUpdate,
    IndustryArticlesUpdate,
)

CSV_DELIMITER = ";"
CATEGORY_COLUMNS: tuple[str, ...] = ("name", "is_active", "is_system")
UNIT_COLUMNS: tuple[str, ...] = ("code", "label", "is_active")


router = APIRouter(prefix="/inventory", tags=["admin-inventory"], dependencies=[Depends(require_admin_key), Depends(get_admin_actor)])


def _category_out(cat: Category) -> CategoryOut:
    return CategoryOut(
        id=str(cat.id),
        name=cat.name,
        is_system=cat.is_system,
        is_active=cat.is_active,
    )


def _item_out_admin(item: Item, category: Category | None) -> ItemOut:
    return ItemOut(
        id=str(item.id),
        sku=item.sku,
        barcode=item.barcode,
        name=item.name,
        description=item.description,
        quantity=item.quantity,
        unit=item.unit,
        is_active=item.is_active,
        category_id=str(item.category_id) if item.category_id else None,
        category_name=category.name if category else None,
        min_stock=item.min_stock,
        max_stock=item.max_stock,
        target_stock=item.target_stock,
        recommended_stock=item.recommended_stock,
        order_mode=item.order_mode,
        is_admin_created=item.is_admin_created,
    )


@router.get("/categories", response_model=list[CategoryOut])
async def admin_list_categories(db: AsyncSession = Depends(get_db)) -> list[CategoryOut]:
    rows = (
        await db.scalars(
            select(Category).where(Category.tenant_id.is_(None)).order_by(Category.is_system.desc(), Category.name.asc())
        )
    ).all()
    return [_category_out(cat) for cat in rows]


@router.post("/categories", response_model=CategoryOut)
async def admin_create_category(payload: CategoryCreate, db: AsyncSession = Depends(get_db)) -> CategoryOut:
    exists = await db.scalar(
        select(Category).where(Category.tenant_id.is_(None), func.lower(Category.name) == func.lower(payload.name))
    )
    if exists:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "category_exists", "message": "Kategorie existiert bereits"}},
        )
    category = Category(
        tenant_id=None,
        name=payload.name.strip(),
        is_system=True,
        is_active=payload.is_active,
    )
    db.add(category)
    await db.commit()
    await db.refresh(category)
    return _category_out(category)


@router.patch("/categories/{category_id}", response_model=CategoryOut)
async def admin_update_category(
    category_id: str,
    payload: CategoryUpdate,
    db: AsyncSession = Depends(get_db),
) -> CategoryOut:
    category = await db.get(Category, category_id)
    if category is None or category.tenant_id is not None:
        raise HTTPException(status_code=404, detail={"error": {"code": "category_not_found", "message": "Kategorie nicht gefunden"}})

    if payload.name:
        conflict = await db.scalar(
            select(Category).where(
                Category.tenant_id.is_(None),
                func.lower(Category.name) == func.lower(payload.name),
                Category.id != category.id,
            )
        )
        if conflict:
            raise HTTPException(
                status_code=400,
                detail={"error": {"code": "category_exists", "message": "Kategorie existiert bereits"}},
            )
        category.name = payload.name.strip()
    if payload.is_active is not None:
        category.is_active = payload.is_active

    await db.commit()
    await db.refresh(category)
    return _category_out(category)


@router.delete("/categories/{category_id}", status_code=204)
async def admin_delete_category(
    category_id: str,
    db: AsyncSession = Depends(get_db),
) -> Response:
    category = await db.get(Category, category_id)
    if category is None or category.tenant_id is not None:
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "category_not_found", "message": "Kategorie nicht gefunden"}},
        )

    # Entkopple Artikel von der Kategorie, damit Delete sauber funktioniert
    items = (await db.scalars(select(Item).where(Item.category_id == category.id))).all()
    for item in items:
        item.category_id = None

    await db.delete(category)
    await db.commit()
    return Response(status_code=204)


@router.post("/categories/import")
async def admin_import_categories(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
) -> dict:
    rows = _rows_from_csv_or_excel(file)
    _ensure_columns(rows, CATEGORY_COLUMNS)

    imported = 0
    updated = 0
    errors: list[dict[str, str]] = []
    for idx, row in enumerate(rows, start=2):
        try:
            name = str(row.get("name") or "").strip()
            if not name:
                raise ValueError("name ist Pflicht")
            is_active = str(row.get("is_active", "true")).strip().lower() in {"1", "true", "yes", "y"}
            is_system = str(row.get("is_system", "true")).strip().lower() in {"1", "true", "yes", "y"}

            existing = await db.scalar(
                select(Category).where(Category.tenant_id.is_(None), func.lower(Category.name) == func.lower(name))
            )
            if existing:
                existing.is_active = is_active
                existing.is_system = is_system
                updated += 1
            else:
                db.add(
                    Category(
                        tenant_id=None,
                        name=name,
                        is_system=is_system,
                        is_active=is_active,
                    )
                )
                imported += 1
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": str(idx), "error": str(exc)})

    await db.commit()
    return {"imported": imported, "updated": updated, "errors": errors}


@router.get("/categories/export")
async def admin_export_categories(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: AsyncSession = Depends(get_db),
):
    import csv

    rows = (await db.scalars(select(Category).where(Category.tenant_id.is_(None)).order_by(Category.name.asc()))).all()
    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Categories"
        ws.append(list(CATEGORY_COLUMNS))
        for cat in rows:
            ws.append([cat.name, cat.is_active, cat.is_system])
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="global_categories.xlsx"'},
        )

    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=CATEGORY_COLUMNS, delimiter=CSV_DELIMITER)
    writer.writeheader()
    for cat in rows:
        writer.writerow({"name": cat.name, "is_active": cat.is_active, "is_system": cat.is_system})
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="global_categories.csv"'},
    )


@router.get("/units", response_model=list[ItemUnitOut])
async def admin_list_units(db: AsyncSession = Depends(get_db)) -> list[ItemUnitOut]:
    rows = (await db.scalars(select(ItemUnit).order_by(ItemUnit.code.asc()))).all()
    return [ItemUnitOut(code=row.code, label=row.label, is_active=row.is_active) for row in rows]


@router.post("/units", response_model=ItemUnitOut)
async def admin_upsert_unit(payload: ItemUnitOut, db: AsyncSession = Depends(get_db)) -> ItemUnitOut:
    unit = await db.get(ItemUnit, payload.code)
    if unit:
        unit.label = payload.label
        unit.is_active = payload.is_active
    else:
        unit = ItemUnit(code=payload.code, label=payload.label, is_active=payload.is_active)
        db.add(unit)
    await db.commit()
    await db.refresh(unit)
    return ItemUnitOut(code=unit.code, label=unit.label, is_active=unit.is_active)


@router.delete("/units/{code}", status_code=204)
async def admin_delete_unit(code: str, db: AsyncSession = Depends(get_db)) -> Response:
    unit = await db.get(ItemUnit, code)
    if unit is None:
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "unit_not_found", "message": "Einheit nicht gefunden"}},
        )

    # Prüfen, ob Einheit genutzt wird
    in_use = await db.scalar(select(func.count()).select_from(Item).where(Item.unit == code))
    if in_use:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "unit_in_use", "message": "Einheit wird von Artikeln verwendet"}},
        )

    await db.delete(unit)
    await db.commit()
    return Response(status_code=204)


@router.post("/units/import")
async def admin_import_units(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
) -> dict:
    rows = _rows_from_csv_or_excel(file)
    _ensure_columns(rows, UNIT_COLUMNS)

    imported = 0
    updated = 0
    errors: list[dict[str, str]] = []
    for idx, row in enumerate(rows, start=2):
        try:
            code = str(row.get("code") or "").strip()
            label = str(row.get("label") or "").strip()
            if not code or not label:
                raise ValueError("code und label sind Pflicht")
            is_active = str(row.get("is_active", "true")).strip().lower() in {"1", "true", "yes", "y"}
            unit = await db.get(ItemUnit, code)
            if unit:
                unit.label = label
                unit.is_active = is_active
                updated += 1
            else:
                db.add(ItemUnit(code=code, label=label, is_active=is_active))
                imported += 1
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": str(idx), "error": str(exc)})

    await db.commit()
    return {"imported": imported, "updated": updated, "errors": errors}


@router.get("/units/export")
async def admin_export_units(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: AsyncSession = Depends(get_db),
):
    import csv

    rows = (await db.scalars(select(ItemUnit).order_by(ItemUnit.code.asc()))).all()
    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Units"
        ws.append(list(UNIT_COLUMNS))
        for unit in rows:
            ws.append([unit.code, unit.label, unit.is_active])
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="global_units.xlsx"'},
        )

    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=UNIT_COLUMNS, delimiter=CSV_DELIMITER)
    writer.writeheader()
    for unit in rows:
        writer.writerow({"code": unit.code, "label": unit.label, "is_active": unit.is_active})
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="global_units.csv"'},
    )


@router.get("/items", response_model=ItemsPage)
async def admin_list_items(
    q: str | None = Query(default=None, description="Suche in SKU, Barcode, Name"),
    category_id: str | None = Query(default=None),
    active: bool | None = Query(default=True),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=200),
    db: AsyncSession = Depends(get_db),
) -> ItemsPage:
    base = select(Item).where(Item.tenant_id.is_(None))
    if active is not None:
        base = base.where(Item.is_active.is_(active))
    if category_id:
        base = base.where(Item.category_id == category_id)
    if q:
        like = f"%{q}%"
        base = base.where(
            (Item.sku.ilike(like)) | (Item.barcode.ilike(like)) | (Item.name.ilike(like))
        )

    total = await db.scalar(select(func.count()).select_from(base.subquery()))
    rows = (
        await db.scalars(base.order_by(Item.name.asc()).offset((page - 1) * page_size).limit(page_size))
    ).all()
    category_ids = {row.category_id for row in rows if row.category_id}
    categories = {}
    if category_ids:
        cat_rows = (await db.scalars(select(Category).where(Category.id.in_(category_ids)))).all()
        categories = {c.id: c for c in cat_rows}

    return ItemsPage(
        items=[_item_out_admin(row, categories.get(row.category_id)) for row in rows],
        total=total or 0,
        page=page,
        page_size=page_size,
    )


@router.post("/items", response_model=ItemOut)
async def admin_create_item(payload: ItemCreate, db: AsyncSession = Depends(get_db)) -> ItemOut:
    normalized_sku = _normalize_sku(payload.sku, prefix_customer=False)
    existing = await db.scalar(select(Item).where(Item.tenant_id.is_(None), Item.sku == normalized_sku))
    if existing:
        raise HTTPException(status_code=400, detail={"error": {"code": "sku_exists", "message": "SKU existiert bereits"}})

    category = await db.get(Category, payload.category_id) if payload.category_id else None
    if category and category.tenant_id is not None:
        raise HTTPException(status_code=400, detail={"error": {"code": "category_forbidden", "message": "Kategorie gehört zu Tenant"}})

    item = Item(
        tenant_id=None,
        sku=normalized_sku,
        barcode=payload.barcode.strip(),
        name=payload.name.strip(),
        description=payload.description or "",
        category_id=category.id if category else None,
        quantity=payload.quantity,
        unit=payload.unit,
        is_active=payload.is_active,
        min_stock=payload.min_stock,
        max_stock=payload.max_stock,
        target_stock=payload.target_stock,
        recommended_stock=payload.recommended_stock,
        order_mode=payload.order_mode,
        is_admin_created=True,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return _item_out_admin(item, category)


@router.patch("/items/{item_id}", response_model=ItemOut)
async def admin_update_item(item_id: str, payload: ItemUpdate, db: AsyncSession = Depends(get_db)) -> ItemOut:
    item = await db.get(Item, item_id)
    if item is None or item.tenant_id is not None:
        raise HTTPException(status_code=404, detail={"error": {"code": "item_not_found", "message": "Artikel nicht gefunden"}})

    category = None
    if payload.category_id is not None:
        category = await db.get(Category, payload.category_id)
        if category and category.tenant_id is not None:
            raise HTTPException(
                status_code=400,
                detail={"error": {"code": "category_forbidden", "message": "Kategorie gehört zu Tenant"}},
            )
        item.category_id = category.id if category else None
    else:
        category = await db.get(Category, item.category_id) if item.category_id else None

    if payload.sku:
        normalized_sku = _normalize_sku(payload.sku, prefix_customer=False)
        exists = await db.scalar(
            select(Item).where(
                Item.tenant_id.is_(None),
                Item.sku == normalized_sku,
                Item.id != item.id,
            )
        )
        if exists:
            raise HTTPException(status_code=400, detail={"error": {"code": "sku_exists", "message": "SKU existiert bereits"}})
        item.sku = normalized_sku

    if payload.barcode is not None:
        item.barcode = payload.barcode.strip()
    if payload.name is not None:
        item.name = payload.name.strip()
    if payload.description is not None:
        item.description = payload.description
    if payload.quantity is not None:
        item.quantity = payload.quantity
    if payload.unit is not None:
        item.unit = payload.unit
    if payload.is_active is not None:
        item.is_active = payload.is_active
    if payload.min_stock is not None:
        item.min_stock = payload.min_stock
    if payload.max_stock is not None:
        item.max_stock = payload.max_stock
    if payload.target_stock is not None:
        item.target_stock = payload.target_stock
    if payload.recommended_stock is not None:
        item.recommended_stock = payload.recommended_stock
    if payload.order_mode is not None:
        item.order_mode = payload.order_mode

    await db.commit()
    await db.refresh(item)
    return _item_out_admin(item, category)


@router.delete("/items/{item_id}", status_code=204)
async def admin_delete_item(item_id: str, db: AsyncSession = Depends(get_db)) -> Response:
    item = await db.get(Item, item_id)
    if item is None or item.tenant_id is not None:
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "item_not_found", "message": "Artikel nicht gefunden"}},
        )

    await db.delete(item)
    await db.commit()
    return Response(status_code=204)


def _rows_from_csv_or_excel(file: UploadFile) -> list[dict]:
    import csv

    filename = (file.filename or "").lower()
    if filename.endswith((".xlsx", ".xls")):
        try:
            wb = load_workbook(BytesIO(file.file.read()))
            ws = wb.active
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=400, detail={"error": {"code": "invalid_excel", "message": str(exc)}})
        data: list[dict] = []
        headers: list[str] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if idx == 1:
                headers = [str(cell).strip() if cell is not None else "" for cell in row]
                continue
            values: dict[str, str] = {}
            for col_idx, header in enumerate(headers):
                values[header] = row[col_idx] if col_idx < len(row) else None
            data.append(values)
        return data

    # CSV mit Semikolon
    content = file.file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail={"error": {"code": "invalid_encoding", "message": "CSV nicht UTF-8"}})

    reader = csv.DictReader(text.splitlines(), delimiter=CSV_DELIMITER)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail={"error": {"code": "invalid_csv", "message": "CSV hat keine Header-Zeile"}})
    return list(reader)


def _ensure_columns(rows: list[dict], required: tuple[str, ...]) -> None:
    if not rows:
        raise HTTPException(status_code=400, detail={"error": {"code": "empty_payload", "message": "Keine Daten gefunden"}})
    headers = {key for key in rows[0].keys() if key is not None}
    missing_columns = [col for col in required if col not in headers]
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "missing_columns", "message": f"Fehlende Spalten: {', '.join(missing_columns)}"}},
        )


@router.post("/items/import")
async def admin_import_items(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    rows = _rows_from_csv_or_excel(file)
    _ensure_columns(rows, CSV_COLUMNS)

    imported = 0
    updated = 0
    errors: list[dict[str, str]] = []

    for idx, row in enumerate(rows, start=2):
        try:
            sku_raw = str(row.get("sku") or "").strip()
            barcode = str(row.get("barcode") or "").strip()
            name = str(row.get("name") or "").strip()
            if not sku_raw or not barcode or not name:
                raise ValueError("sku, barcode und name sind Pflicht")

            normalized_sku = _normalize_sku(sku_raw, prefix_customer=False)
            category_name = str(row.get("category") or "")
            category = None
            if category_name:
                category = await db.scalar(
                    select(Category).where(
                        func.lower(Category.name) == func.lower(category_name),
                        Category.tenant_id.is_(None),
                    )
                )
                if category is None:
                    raise ValueError(f"Kategorie '{category_name}' nicht gefunden")

            quantity = int(row.get("qty") or 0)
            unit = row.get("unit") or "pcs"
            is_active = str(row.get("is_active", "true")).strip().lower() in {"1", "true", "yes", "y"}
            min_stock = int(row.get("min_stock") or 0)
            max_stock = int(row.get("max_stock") or 0)
            target_stock = int(row.get("target_stock") or 0)
            recommended_stock = int(row.get("recommended_stock") or 0)
            order_mode = int(row.get("order_mode") or 0)

            item = await db.scalar(
                select(Item).where(Item.tenant_id.is_(None), Item.sku == normalized_sku)
            )
            if item:
                item.barcode = barcode
                item.name = name
                item.description = row.get("description") or ""
                item.category_id = category.id if category else None
                item.quantity = quantity
                item.unit = unit
                item.is_active = is_active
                item.min_stock = min_stock
                item.max_stock = max_stock
                item.target_stock = target_stock
                item.recommended_stock = recommended_stock
                item.order_mode = order_mode
                item.is_admin_created = True
                updated += 1
            else:
                new_item = Item(
                    tenant_id=None,
                    sku=normalized_sku,
                    barcode=barcode,
                    name=name,
                    description=row.get("description") or "",
                    category_id=category.id if category else None,
                    quantity=quantity,
                    unit=unit,
                    is_active=is_active,
                    min_stock=min_stock,
                    max_stock=max_stock,
                    target_stock=target_stock,
                    recommended_stock=recommended_stock,
                    order_mode=order_mode,
                    is_admin_created=True,
                )
                db.add(new_item)
                imported += 1
        except Exception as e:  # noqa: BLE001
            errors.append({"row": str(idx), "error": str(e)})

    await db.commit()
    return {"imported": imported, "updated": updated, "errors": errors}


@router.get("/items/export")
async def admin_export_items(
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: AsyncSession = Depends(get_db),
):
    import csv

    rows = (
        await db.scalars(
            select(Item)
            .where(Item.tenant_id.is_(None))
            .order_by(Item.name.asc())
        )
    ).all()
    category_ids = {row.category_id for row in rows if row.category_id}
    categories: dict = {}
    if category_ids:
        cat_rows = (await db.scalars(select(Category).where(Category.id.in_(category_ids)))).all()
        categories = {c.id: c for c in cat_rows}

    if format == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "Items"
        ws.append(list(CSV_COLUMNS))
        for item in rows:
            ws.append(
                [
                    item.sku,
                    item.barcode,
                    item.name,
                    item.description,
                    item.quantity,
                    item.unit,
                    item.is_active,
                    categories.get(item.category_id).name if item.category_id else "",
                    item.min_stock,
                    item.max_stock,
                    item.target_stock,
                    item.recommended_stock,
                    item.order_mode,
                ]
            )
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="global_items.xlsx"'},
        )

    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_COLUMNS, delimiter=CSV_DELIMITER)
    writer.writeheader()
    for item in rows:
        writer.writerow(
            {
                "sku": item.sku,
                "barcode": item.barcode,
                "name": item.name,
                "description": item.description,
                "qty": item.quantity,
                "unit": item.unit,
                "is_active": item.is_active,
                "category": categories.get(item.category_id).name if item.category_id else "",
                "min_stock": item.min_stock,
                "max_stock": item.max_stock,
                "target_stock": item.target_stock,
                "recommended_stock": item.recommended_stock,
                "order_mode": item.order_mode,
            }
        )
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="global_items.csv"'},
    )


@router.get("/industries", response_model=list[IndustryOut])
async def admin_list_industries(db: AsyncSession = Depends(get_db)) -> list[IndustryOut]:
    rows = (await db.scalars(select(Industry).order_by(Industry.name.asc()))).all()
    return [
        IndustryOut(
            id=str(entry.id),
            name=entry.name,
            description=entry.description,
            is_active=entry.is_active,
        )
        for entry in rows
    ]


@router.post("/industries", response_model=IndustryOut)
async def admin_create_industry(payload: IndustryCreate, db: AsyncSession = Depends(get_db)) -> IndustryOut:
    existing = await db.scalar(select(Industry).where(func.lower(Industry.name) == func.lower(payload.name)))
    if existing:
        raise HTTPException(status_code=400, detail={"error": {"code": "industry_exists", "message": "Branche existiert bereits"}})
    entry = Industry(
        name=payload.name.strip(),
        description=payload.description or "",
        is_active=payload.is_active,
    )
    db.add(entry)
    await db.commit()
    await db.refresh(entry)
    return IndustryOut(id=str(entry.id), name=entry.name, description=entry.description, is_active=entry.is_active)


@router.patch("/industries/{industry_id}", response_model=IndustryOut)
async def admin_update_industry(
    industry_id: str,
    payload: IndustryUpdate,
    db: AsyncSession = Depends(get_db),
) -> IndustryOut:
    entry = await db.get(Industry, industry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail={"error": {"code": "industry_not_found", "message": "Branche nicht gefunden"}})

    if payload.name:
        conflict = await db.scalar(
            select(Industry).where(func.lower(Industry.name) == func.lower(payload.name), Industry.id != entry.id)
        )
        if conflict:
            raise HTTPException(status_code=400, detail={"error": {"code": "industry_exists", "message": "Branche existiert bereits"}})
        entry.name = payload.name.strip()
    if payload.description is not None:
        entry.description = payload.description
    if payload.is_active is not None:
        entry.is_active = payload.is_active

    await db.commit()
    await db.refresh(entry)
    return IndustryOut(id=str(entry.id), name=entry.name, description=entry.description, is_active=entry.is_active)


@router.delete("/industries/{industry_id}", status_code=204)
async def admin_delete_industry(
    industry_id: str,
    db: AsyncSession = Depends(get_db),
) -> Response:
    entry = await db.get(Industry, industry_id)
    if entry is None:
        raise HTTPException(status_code=404, detail={"error": {"code": "industry_not_found", "message": "Branche nicht gefunden"}})

    await db.delete(entry)
    await db.commit()
    return Response(status_code=204)


@router.get("/industries/{industry_id}/items", response_model=list[ItemOut])
async def admin_list_industry_items(
    industry_id: str = Path(...),
    db: AsyncSession = Depends(get_db),
) -> list[ItemOut]:
    industry = await db.get(Industry, industry_id)
    if industry is None:
        raise HTTPException(status_code=404, detail={"error": {"code": "industry_not_found", "message": "Branche nicht gefunden"}})

    mappings = (
        await db.scalars(
            select(Item)
            .join(IndustryArticle, IndustryArticle.item_id == Item.id)
            .where(IndustryArticle.industry_id == industry.id)
        )
    ).all()
    category_ids = {row.category_id for row in mappings if row.category_id}
    categories = {}
    if category_ids:
        cat_rows = (await db.scalars(select(Category).where(Category.id.in_(category_ids)))).all()
        categories = {c.id: c for c in cat_rows}
    return [_item_out_admin(item, categories.get(item.category_id)) for item in mappings]


@router.put("/industries/{industry_id}/items")
async def admin_set_industry_items(
    industry_id: str,
    payload: IndustryArticlesUpdate,
    db: AsyncSession = Depends(get_db),
) -> dict:
    industry = await db.get(Industry, industry_id)
    if industry is None:
        raise HTTPException(status_code=404, detail={"error": {"code": "industry_not_found", "message": "Branche nicht gefunden"}})

    if payload.item_ids:
        rows = (
            await db.scalars(
                select(Item).where(Item.id.in_(payload.item_ids), Item.tenant_id.is_(None))
            )
        ).all()
        found_ids = {str(row.id) for row in rows}
        missing = [iid for iid in payload.item_ids if iid not in found_ids]
        if missing:
            raise HTTPException(status_code=400, detail={"error": {"code": "item_not_found", "message": f"Unbekannte Artikel: {', '.join(missing)}"}})

    await db.execute(delete(IndustryArticle).where(IndustryArticle.industry_id == industry.id))
    for item_id in payload.item_ids:
        db.add(IndustryArticle(industry_id=industry.id, item_id=item_id))
    await db.commit()
    return {"ok": True, "count": len(payload.item_ids)}
