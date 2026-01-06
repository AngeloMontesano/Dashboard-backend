from __future__ import annotations

import uuid
from datetime import date, datetime, timedelta, timezone
from io import BytesIO, StringIO
from typing import Any, Literal

from fastapi import APIRouter, Depends, File, HTTPException, Path, Query, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from app.core.db import get_db
from app.core.config import settings
from app.core.deps_auth import CurrentUserContext, get_current_user, require_owner_or_admin
from app.core.deps_tenant import get_tenant_context
from app.core.tenant import TenantContext
from app.models.category import Category
from app.models.item import Item
from app.models.item_unit import ItemUnit
from app.models.movement import InventoryMovement
from app.models.order import InventoryOrder, InventoryOrderItem
from app.models.tenant_setting import TenantSetting
from app.modules.inventory.schemas import (
    CategoryCreate,
    CategoryOut,
    CategoryUpdate,
    ItemCreate,
    ItemUnitOut,
    ItemOut,
    ItemUpdate,
    ItemsPage,
    OrderCreate,
    OrderItemOut,
    OrderOut,
    ReportResponse,
    ReportSeries,
    ReportDataPoint,
    ReportKpis,
    ReportTopItem,
    ReorderResponse,
    ReorderItem,
    InventoryBulkUpdateRequest,
    InventoryBulkUpdateResult,
    MovementItemOut,
    MovementOut,
    MovementPayload,
    OrderEmailRequest,
    RecommendedOrderItem,
    RecommendedOrdersResponse,
    SKUExistsResponse,
    TenantSettingsOut,
    TenantSettingsUpdate,
    TenantPingResponse,
    TenantOutPing,
    TestEmailRequest,
    TestEmailResponse,
    MassImportResult,
    EmailSendResponse,
)

router = APIRouter(prefix="/inventory", tags=["inventory"])


@router.get("/ping", response_model=TenantPingResponse)
async def inventory_ping(ctx: TenantContext = Depends(get_tenant_context)) -> TenantPingResponse:
    return TenantPingResponse(
        ok=True,
        tenant=TenantOutPing(
            id=str(ctx.tenant.id),
            slug=ctx.tenant.slug,
            name=ctx.tenant.name,
            is_active=ctx.tenant.is_active,
        ),
    )


# ----------------------
# Kategorien
# ----------------------
@router.get("/categories", response_model=list[CategoryOut])
async def list_categories(
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[CategoryOut]:
    q = select(Category).where(
        or_(
            Category.tenant_id.is_(None),
            Category.tenant_id == ctx.tenant.id,
        ),
    ).order_by(Category.is_system.desc(), Category.name.asc())
    rows = (await db.scalars(q)).all()
    return [
        CategoryOut(
            id=str(c.id),
            name=c.name,
            is_system=c.is_system,
            is_active=c.is_active,
        )
        for c in rows
    ]


@router.get("/units", response_model=list[ItemUnitOut])
async def list_units(
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[ItemUnitOut]:
    rows = (await db.scalars(select(ItemUnit).where(ItemUnit.is_active == True))).all()  # noqa: E712
    return [ItemUnitOut(code=u.code, label=u.label, is_active=u.is_active) for u in rows]


@router.post("/categories", response_model=CategoryOut, dependencies=[Depends(require_owner_or_admin)])
async def create_category(
    payload: CategoryCreate,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> CategoryOut:
    exists = await db.scalar(
        select(Category).where(
            Category.tenant_id == ctx.tenant.id,
            func.lower(Category.name) == func.lower(payload.name),
        )
    )
    if exists:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "category_exists", "message": "Kategorie existiert bereits"}},
        )

    category = Category(
        tenant_id=ctx.tenant.id,
        name=payload.name.strip(),
        is_system=False,
        is_active=payload.is_active,
    )
    db.add(category)
    await db.commit()
    await db.refresh(category)
    return CategoryOut(
        id=str(category.id),
        name=category.name,
        is_system=category.is_system,
        is_active=category.is_active,
    )


@router.patch("/categories/{category_id}", response_model=CategoryOut, dependencies=[Depends(require_owner_or_admin)])
async def update_category(
    category_id: str,
    payload: CategoryUpdate,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> CategoryOut:
    category = await db.get(Category, category_id)
    if category is None or (category.tenant_id and category.tenant_id != ctx.tenant.id):
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "category_not_found", "message": "Kategorie nicht gefunden"}},
        )
    if category.is_system:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "category_system", "message": "Systemkategorien sind schreibgeschützt"}},
        )

    if payload.name:
        conflict = await db.scalar(
            select(Category).where(
                Category.tenant_id == ctx.tenant.id,
                func.lower(Category.name) == func.lower(payload.name),
                Category.id != category.id,
            )
        )
        if conflict:
            raise HTTPException(
                status_code=400,
                detail={"error": {"code": "category_exists", "message": "Kategorie existiert bereits"}},
            )
        category.name = payload.name.strip()
    if payload.is_active is not None:
        category.is_active = payload.is_active

    await db.commit()
    await db.refresh(category)
    return CategoryOut(
        id=str(category.id),
        name=category.name,
        is_system=category.is_system,
        is_active=category.is_active,
    )


# ----------------------
# Items
# ----------------------
def _normalize_sku(sku: str, *, prefix_customer: bool = True) -> str:
    sku = sku.strip()
    if not prefix_customer:
        return sku
    if not sku.lower().startswith("z_"):
        return f"z_{sku}"
    return sku


async def _get_category_or_error(
    *, db: AsyncSession, ctx: TenantContext, category_id: str | None
) -> Category | None:
    if not category_id:
        return None
    category = await db.get(Category, category_id)
    if category is None:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "category_not_found", "message": "Kategorie nicht gefunden"}},
        )
    if category.tenant_id is not None and category.tenant_id != ctx.tenant.id:
        raise HTTPException(
            status_code=403,
            detail={"error": {"code": "category_forbidden", "message": "Kategorie gehört zu anderem Tenant"}},
        )
    if not category.is_active:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "category_inactive", "message": "Kategorie ist inaktiv"}},
        )
    return category


def _item_out(item: Item, category: Category | None) -> ItemOut:
    return ItemOut(
        id=str(item.id),
        sku=item.sku,
        barcode=item.barcode,
        name=item.name,
        description=item.description,
        quantity=item.quantity,
        unit=item.unit,
        is_active=item.is_active,
        category_id=str(item.category_id) if item.category_id else None,
        category_name=category.name if category else None,
        min_stock=item.min_stock,
        max_stock=item.max_stock,
        target_stock=item.target_stock,
        recommended_stock=item.recommended_stock,
        order_mode=item.order_mode,
        is_admin_created=item.is_admin_created,
    )


# ----------------------
# Bestellvorschläge
# ----------------------
@router.get("/orders/recommended", response_model=RecommendedOrdersResponse)
async def list_recommended_orders(
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> RecommendedOrdersResponse:
    rows = (
        await db.scalars(
            select(Item)
            .where(
                Item.tenant_id == ctx.tenant.id,
                Item.is_active.is_(True),
                Item.order_mode.in_([2, 3]),
            )
            .order_by(Item.name.asc())
        )
    ).all()

    category_ids = {row.category_id for row in rows if row.category_id}
    categories: dict[Any, Category] = {}
    if category_ids:
        cat_rows = (await db.scalars(select(Category).where(Category.id.in_(category_ids)))).all()
        categories = {c.id: c for c in cat_rows}

    recommended: list[RecommendedOrderItem] = []
    for item in rows:
        target_level = item.target_stock or item.recommended_stock or item.min_stock
        shortage = max(target_level - item.quantity, 0)
        if shortage == 0 and item.min_stock and item.quantity < item.min_stock:
            shortage = item.min_stock - item.quantity

        if shortage <= 0:
            continue

        recommended_qty = shortage
        if item.max_stock:
            max_allowed = max(item.max_stock - item.quantity, 0)
            recommended_qty = min(recommended_qty, max_allowed) if max_allowed else 0

        if recommended_qty <= 0:
            continue

        category = categories.get(item.category_id)
        recommended.append(
            RecommendedOrderItem(
                item_id=str(item.id),
                sku=item.sku,
                barcode=item.barcode,
                name=item.name,
                quantity=item.quantity,
                min_stock=item.min_stock,
                target_stock=item.target_stock,
                recommended_stock=item.recommended_stock,
                order_mode=item.order_mode,
                recommended_qty=recommended_qty,
                shortage=shortage,
                category_id=str(item.category_id) if item.category_id else None,
                category_name=category.name if category else None,
                unit=item.unit,
            )
        )

    return RecommendedOrdersResponse(items=recommended, total=len(recommended))


@router.get("/items", response_model=ItemsPage)
async def list_items(
    q: str | None = Query(default=None, description="Suche in SKU, Barcode, Name"),
    category_id: str | None = Query(default=None),
    active: bool | None = Query(default=True),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ItemsPage:
    base = select(Item).where(Item.tenant_id == ctx.tenant.id)
    if active is not None:
        base = base.where(Item.is_active.is_(active))
    if category_id:
        base = base.where(Item.category_id == category_id)
    if q:
        like = f"%{q}%"
        base = base.where(
            or_(
                Item.sku.ilike(like),
                Item.barcode.ilike(like),
                Item.name.ilike(like),
            )
        )

    total = await db.scalar(select(func.count()).select_from(base.subquery()))
    rows = (
        await db.scalars(base.order_by(Item.name.asc()).offset((page - 1) * page_size).limit(page_size))
    ).all()

    # Preload categories for output
    category_ids = {row.category_id for row in rows if row.category_id}
    categories: dict[Any, Category] = {}
    if category_ids:
        cat_rows = (
            await db.scalars(select(Category).where(Category.id.in_(category_ids)))
        ).all()
        categories = {c.id: c for c in cat_rows}

    return ItemsPage(
        items=[_item_out(row, categories.get(row.category_id)) for row in rows],
        total=total or 0,
        page=page,
        page_size=page_size,
    )


@router.get("/items/{item_id}", response_model=ItemOut)
async def get_item(
    item_id: str,
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ItemOut:
    item = await db.get(Item, item_id)
    if item is None or item.tenant_id != ctx.tenant.id:
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "item_not_found", "message": "Artikel nicht gefunden"}},
        )
    if item.is_admin_created:
        raise HTTPException(
            status_code=403,
            detail={"error": {"code": "item_readonly", "message": "Artikel wurde durch Admin erstellt und ist schreibgeschützt"}},
        )

    category = None
    if item.category_id:
        category = await db.get(Category, item.category_id)
    return _item_out(item, category)


@router.get("/items/sku/{sku}/exists", response_model=SKUExistsResponse)
async def sku_exists(
    sku: str,
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> SKUExistsResponse:
    normalized = _normalize_sku(sku)
    exists = await db.scalar(
        select(func.count())
        .select_from(Item)
        .where(
            Item.tenant_id == ctx.tenant.id,
            Item.sku == normalized,
        )
    )
    return SKUExistsResponse(exists=bool(exists), normalized_sku=normalized)


@router.post("/items", response_model=ItemOut, dependencies=[Depends(require_owner_or_admin)])
async def create_item(
    payload: ItemCreate,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> ItemOut:
    normalized_sku = _normalize_sku(payload.sku, prefix_customer=True)
    existing = await db.scalar(
        select(Item).where(Item.tenant_id == ctx.tenant.id, Item.sku == normalized_sku)
    )
    if existing:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "sku_exists", "message": "SKU existiert bereits"}},
        )

    category = await _get_category_or_error(db=db, ctx=ctx, category_id=payload.category_id)

    item = Item(
        tenant_id=ctx.tenant.id,
        sku=normalized_sku,
        barcode=payload.barcode.strip(),
        name=payload.name.strip(),
        description=payload.description or "",
        category_id=category.id if category else None,
        quantity=payload.quantity,
        unit=payload.unit,
        is_active=payload.is_active,
        min_stock=payload.min_stock,
        max_stock=payload.max_stock,
        target_stock=payload.target_stock,
        recommended_stock=payload.recommended_stock,
        order_mode=payload.order_mode,
        is_admin_created=False,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return _item_out(item, category)


@router.patch("/items/{item_id}", response_model=ItemOut, dependencies=[Depends(require_owner_or_admin)])
async def update_item(
    item_id: str,
    payload: ItemUpdate,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> ItemOut:
    item = await db.get(Item, item_id)
    if item is None or item.tenant_id != ctx.tenant.id:
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "item_not_found", "message": "Artikel nicht gefunden"}},
        )

    category = None
    if payload.category_id is not None:
        category = await _get_category_or_error(db=db, ctx=ctx, category_id=payload.category_id)
        item.category_id = category.id
    elif payload.category_id is None:
        category = await db.get(Category, item.category_id) if item.category_id else None

    if payload.sku:
        normalized_sku = _normalize_sku(payload.sku, prefix_customer=True)
        exists = await db.scalar(
            select(Item).where(
                Item.tenant_id == ctx.tenant.id,
                Item.sku == normalized_sku,
                Item.id != item.id,
            )
        )
        if exists:
            raise HTTPException(
                status_code=400,
                detail={"error": {"code": "sku_exists", "message": "SKU existiert bereits"}},
            )
        item.sku = normalized_sku

    if payload.barcode is not None:
        item.barcode = payload.barcode.strip()
    if payload.name is not None:
        item.name = payload.name.strip()
    if payload.description is not None:
        item.description = payload.description
    if payload.quantity is not None:
        item.quantity = payload.quantity
    if payload.unit is not None:
        item.unit = payload.unit
    if payload.is_active is not None:
        item.is_active = payload.is_active
    if payload.min_stock is not None:
        item.min_stock = payload.min_stock
    if payload.max_stock is not None:
        item.max_stock = payload.max_stock
    if payload.target_stock is not None:
        item.target_stock = payload.target_stock
    if payload.recommended_stock is not None:
        item.recommended_stock = payload.recommended_stock
    if payload.order_mode is not None:
        item.order_mode = payload.order_mode

    await db.commit()
    await db.refresh(item)

    if item.category_id and category is None:
        category = await db.get(Category, item.category_id)

    return _item_out(item, category)


# ----------------------
# Reporting Helpers
# ----------------------
def _parse_date(value: str, name: str) -> date:
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except Exception:  # noqa: BLE001
        raise HTTPException(status_code=400, detail={"error": {"code": "invalid_date", "message": f"{name} muss YYYY-MM-DD sein"}})


async def _aggregate_consumption(
    *,
    db: AsyncSession,
    ctx: TenantContext,
    start: date,
    end: date,
    mode: Literal["top5", "all", "selected"],
    item_ids: list[str] | None,
    category_id: str | None,
    aggregate: bool | None,
    limit: int | None,
) -> ReportResponse:
    if start > end:
        raise HTTPException(status_code=400, detail={"error": {"code": "invalid_range", "message": "from muss vor to liegen"}})

    start_dt = datetime.combine(start, datetime.min.time()).replace(tzinfo=timezone.utc)
    end_dt = datetime.combine(end + timedelta(days=1), datetime.min.time()).replace(tzinfo=timezone.utc)

    period_col = func.date_trunc("month", InventoryMovement.created_at).label("period")

    base = (
        select(
            period_col,
            Item.id.label("item_id"),
            Item.name.label("item_name"),
            Item.sku.label("item_sku"),
            Item.category_id.label("category_id"),
            func.sum(InventoryMovement.qty).label("qty_sum"),
        )
        .join(Item, InventoryMovement.item_id == Item.id)
        .where(
            InventoryMovement.tenant_id == ctx.tenant.id,
            InventoryMovement.type == "OUT",
            InventoryMovement.created_at >= start_dt,
            InventoryMovement.created_at < end_dt,
        )
    )
    if category_id:
        base = base.where(Item.category_id == category_id)
    if item_ids:
        base = base.where(Item.id.in_(item_ids))

    grouped = (
        await db.execute(
            base.group_by(
                period_col,
                Item.id,
                Item.name,
                Item.sku,
                Item.category_id,
            ).order_by(period_col)
        )
    ).all()

    series_map: dict[str, dict[str, Any]] = {}
    months: set[str] = set()

    for row in grouped:
        period: datetime = row.period
        month_key = period.strftime("%Y-%m")
        months.add(month_key)
        item_id = str(row.item_id)
        entry = series_map.setdefault(
            item_id,
            {"name": row.item_name, "category_id": row.category_id, "months": {}},
        )
        entry["months"][month_key] = float(row.qty_sum or 0.0)

    month_list = sorted(months)

    def _series_from_entry(item_id: str, entry: dict[str, Any]) -> ReportSeries:
        data = [
            ReportDataPoint(
                period=month,
                value=entry["months"].get(month, 0.0),
                item_id=item_id,
                item_name=entry["name"],
            )
            for month in month_list
        ]
        return ReportSeries(label=entry["name"], itemId=item_id, data=data)

    totals_by_item = [
        (item_id, sum(values["months"].values()), values)
        for item_id, values in series_map.items()
    ]
    totals_by_item.sort(key=lambda x: x[1], reverse=True)

    series: list[ReportSeries] = []

    if mode == "selected":
        selected_ids = item_ids or []
        for item_id in selected_ids:
            entry = series_map.get(item_id)
            if entry:
                series.append(_series_from_entry(item_id, entry))
    elif mode == "top5":
        take = limit or 5
        for item_id, _total, entry in totals_by_item[:take]:
            series.append(_series_from_entry(item_id, entry))
    else:  # all
        if aggregate is False:
            for item_id, _total, entry in totals_by_item:
                series.append(_series_from_entry(item_id, entry))
        else:
            agg_per_month: dict[str, float] = {m: 0.0 for m in month_list}
            for _item_id, _total, entry in totals_by_item:
                for month, value in entry["months"].items():
                    agg_per_month[month] = agg_per_month.get(month, 0.0) + value
            data = [
                ReportDataPoint(period=month, value=agg_per_month.get(month, 0.0))
                for month in month_list
            ]
            series.append(ReportSeries(label="Alle Artikel", data=data))

    total_consumption = sum(total for _id, total, _entry in totals_by_item)
    average_per_month = total_consumption / len(month_list) if month_list else 0.0
    top_item = totals_by_item[0] if totals_by_item else None
    kpis = ReportKpis(
        totalConsumption=total_consumption,
        averagePerMonth=average_per_month,
        months=month_list,
        topItem=ReportTopItem(
            id=top_item[0],
            name=top_item[2]["name"],
            quantity=top_item[1],
        )
        if top_item
        else None,
    )
    return ReportResponse(series=series, kpis=kpis)


# ----------------------
# Reporting (Verbrauch)
# ----------------------
@router.get("/report", response_model=ReportResponse)
async def get_report(
    from_: str = Query(..., alias="from", description="Startdatum (YYYY-MM-DD)"),
    to: str = Query(..., description="Enddatum (YYYY-MM-DD)"),
    mode: Literal["top5", "all", "selected"] = Query(..., description="Aggregationsmodus"),
    item_ids: list[str] | None = Query(default=None),
    category_id: str | None = Query(default=None),
    aggregate: bool | None = Query(default=True),
    limit: int | None = Query(default=5),
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ReportResponse:
    start_date = _parse_date(from_, "from")
    end_date = _parse_date(to, "to")
    return await _aggregate_consumption(
        db=db,
        ctx=ctx,
        start=start_date,
        end=end_date,
        mode=mode,
        item_ids=item_ids,
        category_id=category_id,
        aggregate=aggregate,
        limit=limit,
    )


@router.get("/reports/consumption", response_model=ReportResponse)
async def get_report_consumption(
    from_: str = Query(..., alias="from", description="Startdatum (YYYY-MM-DD)"),
    to: str = Query(..., description="Enddatum (YYYY-MM-DD)"),
    mode: Literal["top5", "all", "selected"] = Query(..., description="Aggregationsmodus"),
    item_ids: list[str] | None = Query(default=None),
    category_id: str | None = Query(default=None),
    aggregate: bool | None = Query(default=True),
    limit: int | None = Query(default=5),
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ReportResponse:
    return await get_report(
        from_=from_,
        to=to,
        mode=mode,
        item_ids=item_ids,
        category_id=category_id,
        aggregate=aggregate,
        limit=limit,
        ctx=ctx,
        user_ctx=user_ctx,
        db=db,
    )


@router.get("/reports/export/{format}")
async def export_report(
    format: str = Path(..., description="csv oder excel"),
    from_: str = Query(..., alias="from", description="Startdatum (YYYY-MM-DD)"),
    to: str = Query(..., description="Enddatum (YYYY-MM-DD)"),
    mode: Literal["top5", "all", "selected"] = Query(..., description="Aggregationsmodus"),
    item_ids: list[str] | None = Query(default=None),
    category_id: str | None = Query(default=None),
    aggregate: bool | None = Query(default=True),
    limit: int | None = Query(default=5),
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if format not in {"csv", "excel"}:
        raise HTTPException(status_code=400, detail={"error": {"code": "invalid_format", "message": "Format muss csv oder excel sein"}})

    report = await get_report(
        from_=from_,
        to=to,
        mode=mode,
        item_ids=item_ids,
        category_id=category_id,
        aggregate=aggregate,
        limit=limit,
        ctx=ctx,
        user_ctx=user_ctx,
        db=db,
    )

    rows: list[tuple[str, str, float]] = []
    for serie in report.series:
        for point in serie.data:
            rows.append((serie.label, point.period, point.value))

    if format == "csv":
        import csv

        buf = StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Artikel", "Monat", "Verbrauch"])
        for artikel, month, value in rows:
            writer.writerow([artikel, month, value])
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="verbrauch.csv"'},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Verbrauch"
    ws.append(["Artikel", "Monat", "Verbrauch"])
    for artikel, month, value in rows:
        ws.append([artikel, month, value])
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="verbrauch.xlsx"'},
    )


# ----------------------
# Bestellungen (Offen/Erledigt)
# ----------------------
def _generate_order_number() -> str:
    return f"ORDER-{uuid.uuid4().hex[:8].upper()}"


def _order_item_to_out(order_item: InventoryOrderItem, item_map: dict[Any, Item]) -> OrderItemOut:
    item = item_map.get(str(order_item.item_id))
    return OrderItemOut(
        id=str(order_item.id),
        item_id=str(order_item.item_id),
        quantity=order_item.quantity,
        note=order_item.note,
        item_name=item.name if item else None,
        sku=item.sku if item else None,
        barcode=item.barcode if item else None,
        category_id=str(item.category_id) if item and item.category_id else None,
    )


def _order_to_out(order: InventoryOrder, item_map: dict[Any, Item]) -> OrderOut:
    return OrderOut(
        id=str(order.id),
        number=order.number,
        status=order.status,  # type: ignore[arg-type]
        note=order.note,
        created_at=order.created_at,
        completed_at=order.completed_at,
        canceled_at=order.canceled_at,
        items=[_order_item_to_out(oi, item_map) for oi in order.items],
    )


async def _get_order_or_404(
    *, order_id: str, ctx: TenantContext, db: AsyncSession, load_items: bool = False
) -> InventoryOrder:
    try:
        order_uuid = uuid.UUID(str(order_id))
    except (TypeError, ValueError):
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "order_not_found", "message": "Bestellung nicht gefunden"}},
        )
    stmt = select(InventoryOrder).where(
        InventoryOrder.tenant_id == ctx.tenant.id,
        InventoryOrder.id == order_uuid,
    )
    if load_items:
        stmt = stmt.options(selectinload(InventoryOrder.items))
    order = await db.scalar(stmt)
    if order is None:
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "order_not_found", "message": "Bestellung nicht gefunden"}},
        )
    return order


async def _items_by_ids(
    *, db: AsyncSession, ctx: TenantContext, item_ids: set[Any]
) -> dict[str, Item]:
    if not item_ids:
        return {}
    normalized_ids: set[uuid.UUID] = set()
    try:
        for raw in item_ids:
            normalized_ids.add(uuid.UUID(str(raw)))
    except (TypeError, ValueError):
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "invalid_item_id", "message": "Ungültige Item-ID"}},
        )
    rows = (
        await db.scalars(
            select(Item).where(
                Item.tenant_id == ctx.tenant.id,
                Item.id.in_(normalized_ids),
            )
        )
    ).all()
    return {str(row.id): row for row in rows}


@router.get("/orders", response_model=list[OrderOut])
async def list_orders(
    status: Literal["OPEN", "COMPLETED", "CANCELED"] | None = Query(default=None),
    limit: int = Query(default=100, ge=1, le=500),
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[OrderOut]:
    stmt = (
        select(InventoryOrder)
        .where(InventoryOrder.tenant_id == ctx.tenant.id)
        .options(selectinload(InventoryOrder.items))
    )
    if status:
        stmt = stmt.where(InventoryOrder.status == status)
    stmt = stmt.order_by(InventoryOrder.created_at.desc()).limit(limit)

    orders = (await db.scalars(stmt)).all()
    item_ids = {str(oi.item_id) for order in orders for oi in order.items}
    item_map = await _items_by_ids(db=db, ctx=ctx, item_ids=item_ids)
    return [_order_to_out(order, item_map) for order in orders]


@router.post("/orders", response_model=OrderOut, status_code=201, dependencies=[Depends(require_owner_or_admin)])
async def create_order(
    payload: OrderCreate,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> OrderOut:
    if not payload.items:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "items_required", "message": "Mindestens eine Position erforderlich"}},
        )

    item_ids = {entry.item_id for entry in payload.items}
    item_map = await _items_by_ids(db=db, ctx=ctx, item_ids=item_ids)
    missing = [iid for iid in item_ids if iid not in item_map]
    if missing:
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "item_not_found", "message": f"Artikel nicht gefunden: {', '.join(missing)}"}},
        )

    order = InventoryOrder(
        tenant_id=ctx.tenant.id,
        number=_generate_order_number(),
        status="OPEN",
        note=payload.note,
    )
    for entry in payload.items:
        order.items.append(
            InventoryOrderItem(
                tenant_id=ctx.tenant.id,
                item_id=item_map[entry.item_id].id,
                quantity=entry.quantity,
                note=entry.note,
            )
        )

    db.add(order)
    await db.flush()
    await db.commit()

    order = await _get_order_or_404(order_id=str(order.id), ctx=ctx, db=db, load_items=True)
    return _order_to_out(order, item_map)


@router.get("/orders/{order_id}", response_model=OrderOut)
async def get_order(
    order_id: str,
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> OrderOut:
    order = await _get_order_or_404(order_id=order_id, ctx=ctx, db=db, load_items=True)
    item_ids = {str(oi.item_id) for oi in order.items}
    item_map = await _items_by_ids(db=db, ctx=ctx, item_ids=item_ids)
    return _order_to_out(order, item_map)


@router.post("/orders/{order_id}/complete", response_model=OrderOut, dependencies=[Depends(require_owner_or_admin)])
async def complete_order(
    order_id: str,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> OrderOut:
    order = await _get_order_or_404(order_id=order_id, ctx=ctx, db=db, load_items=True)
    if order.status != "OPEN":
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "order_not_open", "message": "Nur offene Bestellungen können erledigt werden"}},
        )

    item_ids = {str(oi.item_id) for oi in order.items}
    item_map = await _items_by_ids(db=db, ctx=ctx, item_ids=item_ids)

    now = datetime.now(timezone.utc)
    order.status = "COMPLETED"
    order.completed_at = now

    for order_item in order.items:
        item = item_map.get(str(order_item.item_id))
        if item is None:
            raise HTTPException(
                status_code=400,
                detail={"error": {"code": "order_item_missing", "message": "Artikel für Bestellung fehlt"}},
            )
        item.quantity += order_item.quantity
        movement = InventoryMovement(
            tenant_id=ctx.tenant.id,
            item_id=item.id,
            client_tx_id=f"order-{order.id}-{order_item.id}",
            type="IN",
            barcode=item.barcode,
            qty=order_item.quantity,
            note=f"Order {order.number}",
            created_at=now,
        )
        db.add(movement)

    await db.commit()
    await db.refresh(order)
    return _order_to_out(order, item_map)


@router.post("/orders/{order_id}/cancel", response_model=OrderOut, dependencies=[Depends(require_owner_or_admin)])
async def cancel_order(
    order_id: str,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> OrderOut:
    order = await _get_order_or_404(order_id=order_id, ctx=ctx, db=db, load_items=True)
    if order.status != "OPEN":
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "order_not_open", "message": "Nur offene Bestellungen können storniert werden"}},
        )

    order.status = "CANCELED"
    order.canceled_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(order)

    item_ids = {oi.item_id for oi in order.items}
    item_map = await _items_by_ids(db=db, ctx=ctx, item_ids=item_ids)
    return _order_to_out(order, item_map)


def _send_email_message(*, to_email: str, subject: str, body: str) -> EmailSendResponse:
    if not settings.SMTP_HOST or not settings.SMTP_PORT or not settings.SMTP_FROM:
        return EmailSendResponse(ok=False, error="SMTP Konfiguration fehlt")

    import smtplib
    from email.message import EmailMessage

    message = EmailMessage()
    message["Subject"] = subject
    message["From"] = settings.SMTP_FROM
    message["To"] = to_email
    message.set_content(body)

    try:
        with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT, timeout=10) as smtp:
            if settings.SMTP_USER and settings.SMTP_PASSWORD:
                smtp.starttls()
                smtp.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
            smtp.send_message(message)
        return EmailSendResponse(ok=True, error=None)
    except Exception as e:  # noqa: BLE001
        return EmailSendResponse(ok=False, error=str(e))


def _format_order_email(
    order: InventoryOrder,
    item_map: dict[str, Item],
    recipient: str,
    note: str | None,
) -> tuple[str, str]:
    lines = [
        f"Empfänger: {recipient}",
        f"Bestellung: {order.number}",
        f"Status: {order.status}",
        f"Angelegt: {order.created_at}",
    ]
    if order.note:
        lines.append(f"Bestell-Notiz: {order.note}")
    if note:
        lines.append(f"Zusätzliche Notiz: {note}")
    lines.append("")
    lines.append("Positionen:")
    for oi in order.items:
        item = item_map.get(str(oi.item_id))
        item_name = item.name if item else "(unbekannt)"
        lines.append(
            f"- {item_name} (Menge: {oi.quantity}, SKU: {item.sku if item else '-'}, Barcode: {item.barcode if item else '-'})"
        )
    subject = f"Bestellung {order.number}"
    body = "\n".join(lines)
    return subject, body


def _build_order_pdf(order: InventoryOrder, item_map: dict[str, Item]) -> bytes:
    buf = BytesIO()
    pdf = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    y = height - 40

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, y, f"Bestellung {order.number}")
    y -= 20
    pdf.setFont("Helvetica", 10)
    pdf.drawString(40, y, f"Status: {order.status}")
    y -= 15
    if order.note:
        pdf.drawString(40, y, f"Notiz: {order.note}")
        y -= 15
    pdf.drawString(40, y, f"Angelegt: {order.created_at}")
    y -= 25

    pdf.setFont("Helvetica-Bold", 11)
    pdf.drawString(40, y, "Positionen:")
    y -= 18
    pdf.setFont("Helvetica", 10)

    headers = ["Artikel", "Menge", "SKU", "Barcode"]
    col_x = [40, 250, 320, 420]
    for idx, header in enumerate(headers):
        pdf.drawString(col_x[idx], y, header)
    y -= 12
    pdf.line(40, y, width - 40, y)
    y -= 14

    for oi in order.items:
        if y < 60:
            pdf.showPage()
            y = height - 40
            pdf.setFont("Helvetica", 10)
        item = item_map.get(str(oi.item_id))
        name = item.name if item else "(unbekannt)"
        sku = item.sku if item else "-"
        barcode = item.barcode if item else "-"
        values = [name, str(oi.quantity), sku, barcode]
        for idx, value in enumerate(values):
            pdf.drawString(col_x[idx], y, value)
        y -= 14

    pdf.showPage()
    pdf.save()
    buf.seek(0)
    return buf.getvalue()


@router.post("/orders/{order_id}/email", response_model=EmailSendResponse, dependencies=[Depends(require_owner_or_admin)])
async def send_order_email(
    order_id: str,
    payload: OrderEmailRequest,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> EmailSendResponse:
    order = await _get_order_or_404(order_id=order_id, ctx=ctx, db=db, load_items=True)
    item_ids = {str(oi.item_id) for oi in order.items}
    item_map = await _items_by_ids(db=db, ctx=ctx, item_ids=item_ids)
    settings_obj = await _get_or_create_settings(ctx=ctx, db=db)

    recipient = payload.email or settings_obj.order_email or settings_obj.contact_email
    if not recipient:
        return EmailSendResponse(ok=False, error="Kein Empfänger konfiguriert")

    subject, body = _format_order_email(order, item_map, recipient, payload.note)
    return _send_email_message(to_email=recipient, subject=subject, body=body)


@router.get("/orders/{order_id}/pdf")
async def get_order_pdf(
    order_id: str,
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    order = await _get_order_or_404(order_id=order_id, ctx=ctx, db=db, load_items=True)
    item_ids = {str(oi.item_id) for oi in order.items}
    item_map = await _items_by_ids(db=db, ctx=ctx, item_ids=item_ids)

    pdf_bytes = _build_order_pdf(order, item_map)
    filename = f"order-{order.number}.pdf"
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ----------------------
# Einstellungen (Tenant)
# ----------------------
MASS_EXPORT_COLUMNS = [
    "Artikel-ID",
    "Name",
    "Barcode",
    "Kategorie",
    "Soll",
    "Min",
    "Bestand",
    "Haltbarkeit",
    "Charge",
    "Kunden-ID",
    "Lagerort",
    "Preis",
    "Beschreibung",
    "Letzte Änderung",
    "Letzte Bestellung",
    "Bestellt",
]


def _settings_to_out(settings: TenantSetting) -> TenantSettingsOut:
    return TenantSettingsOut(
        id=str(settings.id),
        company_name=settings.company_name,
        contact_email=settings.contact_email,
        order_email=settings.order_email,
        auto_order_enabled=settings.auto_order_enabled,
        auto_order_min=settings.auto_order_min,
        export_format=settings.export_format,
        address=settings.address,
        address_postal_code=settings.address_postal_code,
        address_city=settings.address_city,
        phone=settings.phone,
        contact_name=settings.contact_name,
        branch_number=settings.branch_number,
        tax_number=settings.tax_number,
        industry_id=str(settings.industry_id) if settings.industry_id else None,
    )


async def _get_or_create_settings(*, ctx: TenantContext, db: AsyncSession) -> TenantSetting:
    settings = await db.scalar(
        select(TenantSetting).where(TenantSetting.tenant_id == ctx.tenant.id)
    )
    if settings:
        return settings
    settings = TenantSetting(
        tenant_id=ctx.tenant.id,
        company_name="",
        contact_email="",
        order_email="",
        auto_order_enabled=False,
        auto_order_min=0,
        export_format="xlsx",
        address="",
        address_postal_code="",
        address_city="",
        phone="",
        contact_name="",
        branch_number="",
        tax_number="",
        industry_id=None,
    )
    db.add(settings)
    await db.commit()
    await db.refresh(settings)
    return settings


@router.get("/settings", response_model=TenantSettingsOut)
async def get_tenant_settings(
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> TenantSettingsOut:
    settings = await _get_or_create_settings(ctx=ctx, db=db)
    return _settings_to_out(settings)


@router.put("/settings", response_model=TenantSettingsOut, dependencies=[Depends(require_owner_or_admin)])
async def update_tenant_settings(
    payload: TenantSettingsUpdate,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> TenantSettingsOut:
    settings = await _get_or_create_settings(ctx=ctx, db=db)
    settings.company_name = payload.company_name.strip()
    settings.contact_email = payload.contact_email.strip()
    settings.order_email = payload.order_email.strip()
    settings.auto_order_enabled = payload.auto_order_enabled
    settings.auto_order_min = payload.auto_order_min
    settings.export_format = payload.export_format.strip() or "xlsx"
    settings.address = payload.address.strip()
    settings.address_postal_code = payload.address_postal_code.strip()
    settings.address_city = payload.address_city.strip()
    settings.phone = payload.phone.strip()
    settings.contact_name = payload.contact_name.strip()
    settings.branch_number = payload.branch_number.strip()
    settings.tax_number = payload.tax_number.strip()
    settings.industry_id = payload.industry_id

    await db.commit()
    await db.refresh(settings)
    return _settings_to_out(settings)


@router.get("/settings/export", dependencies=[Depends(require_owner_or_admin)])
async def export_settings_inventory(
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
):
    rows = (
        await db.execute(
            select(Item, Category)
            .outerjoin(Category, Category.id == Item.category_id)
            .where(Item.tenant_id == ctx.tenant.id)
            .order_by(Item.name.asc())
        )
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(MASS_EXPORT_COLUMNS)

    for item, category in rows:
        ws.append(
            [
                item.sku,
                item.name,
                item.barcode,
                category.name if category else "",
                item.target_stock,
                item.min_stock,
                item.quantity,
                "",
                "",
                str(ctx.tenant.id),
                "",
                "",
                item.description or "",
                "",
                "",
                "",
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="settings_export.xlsx"'},
    )


@router.post("/settings/import", response_model=MassImportResult, dependencies=[Depends(require_owner_or_admin)])
async def import_settings_inventory(
    file: UploadFile = File(...),
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> MassImportResult:
    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content))
    except Exception:  # noqa: BLE001
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "invalid_excel", "message": "Datei konnte nicht gelesen werden"}},
        )
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    missing = [col for col in MASS_EXPORT_COLUMNS if col not in headers]
    if missing:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "missing_columns", "message": f"Fehlende Spalten: {', '.join(missing)}"}},
        )
    idx = {name: headers.index(name) for name in headers}

    imported = 0
    updated = 0
    errors: list[dict[str, str]] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            sku_raw = str(row[idx["Artikel-ID"]].value or "").strip()
            name = str(row[idx["Name"]].value or "").strip()
            barcode = str(row[idx["Barcode"]].value or "").strip()
            if not sku_raw or not name or not barcode:
                raise ValueError("Artikel-ID, Name und Barcode sind Pflicht")

            category_name = str(row[idx["Kategorie"]].value or "").strip()
            category = await _category_by_name(db=db, ctx=ctx, name=category_name) if category_name else None
            if category_name and category is None:
                raise ValueError(f"Kategorie '{category_name}' nicht gefunden")

            target_stock = int(row[idx["Soll"]].value or 0)
            min_stock = int(row[idx["Min"]].value or 0)
            quantity = int(row[idx["Bestand"]].value or 0)
            description = str(row[idx["Beschreibung"]].value or "").strip()

            normalized_sku = _normalize_sku(sku_raw)

            item = await db.scalar(select(Item).where(Item.tenant_id == ctx.tenant.id, Item.sku == normalized_sku))
            if item:
                item.name = name
                item.barcode = barcode
                item.category_id = category.id if category else None
                item.target_stock = target_stock
                item.min_stock = min_stock
                item.quantity = quantity
                item.description = description
                updated += 1
            else:
                new_item = Item(
                    tenant_id=ctx.tenant.id,
                    sku=normalized_sku,
                    barcode=barcode,
                    name=name,
                    category_id=category.id if category else None,
                    target_stock=target_stock,
                    min_stock=min_stock,
                    quantity=quantity,
                    description=description,
                )
                db.add(new_item)
                imported += 1
        except Exception as e:  # noqa: BLE001
            errors.append({"row": str(row_num), "error": str(e)})

    await db.commit()
    return MassImportResult(imported=imported, updated=updated, errors=errors)


@router.post("/settings/test-email", response_model=TestEmailResponse, dependencies=[Depends(require_owner_or_admin)])
async def send_test_email(
    payload: TestEmailRequest,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> TestEmailResponse:
    """
    Sendet eine Test-E-Mail an die angegebene Adresse, nutzt SMTP-Konfiguration aus Settings.
    """
    if not settings.SMTP_HOST or not settings.SMTP_PORT or not settings.SMTP_FROM:
        return TestEmailResponse(ok=False, error="SMTP Konfiguration fehlt")

    import smtplib
    from email.message import EmailMessage

    message = EmailMessage()
    message["Subject"] = "Test E-Mail Lagerverwaltung"
    message["From"] = settings.SMTP_FROM
    message["To"] = payload.email
    message.set_content(f"Test E-Mail für Tenant {ctx.tenant.name} ({ctx.tenant.slug})")

    try:
        with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT, timeout=10) as smtp:
            if settings.SMTP_USER and settings.SMTP_PASSWORD:
                smtp.starttls()
                smtp.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
            smtp.send_message(message)
        return TestEmailResponse(ok=True, error=None)
    except Exception as e:  # noqa: BLE001
        return TestEmailResponse(ok=False, error=str(e))


# ----------------------
# Bestellwürdig (Empfehlungen)
# ----------------------
@router.get("/orders/recommended", response_model=ReorderResponse)
async def get_reorder_recommendations(
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ReorderResponse:
    diff_expr = Item.target_stock - Item.quantity
    q = select(Item).where(
        Item.tenant_id == ctx.tenant.id,
        Item.is_active.is_(True),
        Item.target_stock > 0,
        Item.order_mode.in_([1, 2, 3]),
        Item.quantity < Item.target_stock,
    ).order_by(diff_expr.desc())

    items = (await db.scalars(q)).all()
    return ReorderResponse(
        items=[
            ReorderItem(
                id=str(item.id),
                name=item.name,
                sku=item.sku,
                barcode=item.barcode,
                category_id=str(item.category_id) if item.category_id else None,
                quantity=item.quantity,
                target_stock=item.target_stock,
                min_stock=item.min_stock,
                recommended_qty=max(
                    item.target_stock - item.quantity,
                    item.min_stock - item.quantity,
                    1,
                ),
            )
            for item in items
        ]
    )


# ----------------------
# Bewegungen (Bestandsbuchungen)
# ----------------------
@router.get("/movements", response_model=list[MovementOut])
async def list_movements(
    start: datetime | None = Query(default=None, description="Startzeitpunkt (inklusive, UTC oder lokal ISO8601)"),
    end: datetime | None = Query(default=None, description="Endzeitpunkt (inklusive, UTC oder lokal ISO8601)"),
    movement_type: Literal["IN", "OUT"] | None = Query(default=None, alias="type"),
    category_id: str | None = Query(default=None),
    item_ids: list[str] | None = Query(default=None, alias="item_ids"),
    limit: int = Query(default=200, ge=1, le=1000, description="Maximale Anzahl zurückgegebener Bewegungen"),
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> list[MovementOut]:
    q = (
        select(InventoryMovement, Item, Category)
        .join(Item, InventoryMovement.item_id == Item.id)
        .outerjoin(Category, Category.id == Item.category_id)
        .where(InventoryMovement.tenant_id == ctx.tenant.id)
    )
    if start:
        q = q.where(InventoryMovement.created_at >= start)
    if end:
        q = q.where(InventoryMovement.created_at <= end)
    if movement_type:
        q = q.where(InventoryMovement.type == movement_type)
    if category_id:
        q = q.where(Item.category_id == category_id)
    if item_ids:
        q = q.where(Item.id.in_(item_ids))

    rows = (
        await db.execute(
            q.order_by(InventoryMovement.created_at.desc()).limit(limit)
        )
    ).all()

    def _movement_to_out(movement: InventoryMovement, item: Item | None) -> MovementOut:
        return MovementOut(
            id=str(movement.id),
            item_id=str(movement.item_id),
            item_name=item.name if item else None,
            category_id=str(item.category_id) if item and item.category_id else None,
            type=movement.type,
            barcode=movement.barcode,
            qty=movement.qty,
            note=movement.note,
            created_at=movement.created_at,
            item=MovementItemOut(
                id=str(item.id),
                sku=item.sku,
                barcode=item.barcode,
                name=item.name,
                category_id=str(item.category_id) if item.category_id else None,
            )
            if item
            else None,
        )

    return [
        _movement_to_out(movement=row[0], item=row[1])
        for row in rows
    ]


@router.post("/movements", dependencies=[Depends(require_owner_or_admin)])
async def create_movement(
    payload: MovementPayload,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
):
    item = await db.scalar(
        select(Item).where(Item.tenant_id == ctx.tenant.id, Item.barcode == payload.barcode.strip())
    )
    if item is None:
        raise HTTPException(
            status_code=404,
            detail={"error": {"code": "item_not_found", "message": "Artikel zum Barcode nicht gefunden"}},
        )

    existing = await db.scalar(
        select(InventoryMovement).where(
            InventoryMovement.tenant_id == ctx.tenant.id,
            InventoryMovement.client_tx_id == payload.client_tx_id,
        )
    )
    if existing:
        category = await db.get(Category, item.category_id) if item.category_id else None
        return {"ok": True, "item": _item_out(item, category), "movement_id": str(existing.id), "duplicate": True}

    delta = payload.qty if payload.type == "IN" else -payload.qty
    new_qty = item.quantity + delta
    if new_qty < 0:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "qty_negative", "message": "Bestand würde negativ werden"}},
        )

    item.quantity = new_qty
    movement = InventoryMovement(
        tenant_id=ctx.tenant.id,
        item_id=item.id,
        client_tx_id=payload.client_tx_id,
        type=payload.type,
        barcode=item.barcode,
        qty=payload.qty,
        note=payload.note,
        created_at=payload.created_at or datetime.now(timezone.utc),
    )
    db.add(movement)
    await db.commit()
    await db.refresh(item)
    await db.refresh(movement)

    category = await db.get(Category, item.category_id) if item.category_id else None
    return {"ok": True, "item": _item_out(item, category), "movement_id": str(movement.id), "duplicate": False}


# ----------------------
# Inventur (Bulk-Update & Export)
# ----------------------
@router.post("/inventory/bulk", response_model=InventoryBulkUpdateResult, dependencies=[Depends(require_owner_or_admin)])
async def bulk_update_inventory(
    payload: InventoryBulkUpdateRequest,
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> InventoryBulkUpdateResult:
    errors: list[str] = []
    updated = 0

    item_ids = [u.item_id for u in payload.updates]
    if not item_ids:
        return InventoryBulkUpdateResult(updated=0, errors=[])

    items = (
        await db.scalars(select(Item).where(Item.tenant_id == ctx.tenant.id, Item.id.in_(item_ids)))
    ).all()
    item_map = {str(item.id): item for item in items}

    for entry in payload.updates:
        item = item_map.get(entry.item_id)
        if not item:
            errors.append(f"Item {entry.item_id} not found for tenant")
            continue
        item.quantity = entry.quantity
        updated += 1

    await db.commit()
    return InventoryBulkUpdateResult(updated=updated, errors=errors)


@router.get("/inventory/export")
async def export_inventory(
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    rows = (
        await db.execute(
            select(Item, Category)
            .outerjoin(Category, Category.id == Item.category_id)
            .where(Item.tenant_id == ctx.tenant.id)
            .order_by(Item.name.asc())
        )
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventur"

    headers = ["Artikel-ID", "Name", "Barcode", "Kategorie", "Soll", "Min", "Bestand"]
    ws.append(headers)

    for item, category in rows:
        ws.append(
            [
                item.sku,
                item.name,
                item.barcode,
                category.name if category else "",
                item.target_stock,
                item.min_stock,
                item.quantity,
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "inventur.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ----------------------
# CSV Import/Export (Basis)
# ----------------------
CSV_COLUMNS: tuple[str, ...] = (
    "sku",
    "barcode",
    "name",
    "description",
    "qty",
    "unit",
    "is_active",
    "category",
    "min_stock",
    "max_stock",
    "target_stock",
    "recommended_stock",
    "order_mode",
)


def _parse_bool(value: str) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


async def _category_by_name(
    *, db: AsyncSession, ctx: TenantContext, name: str | None
) -> Category | None:
    if not name:
        return None
    return await db.scalar(
        select(Category).where(
            func.lower(Category.name) == func.lower(name.strip()),
            or_(Category.tenant_id == ctx.tenant.id, Category.tenant_id.is_(None)),
        )
    )


def _validate_order_mode(value: str) -> int:
    try:
        mode = int(value)
    except Exception:
        raise ValueError("order_mode muss 0,1,2 oder 3 sein")
    if mode not in (0, 1, 2, 3):
        raise ValueError("order_mode muss 0,1,2 oder 3 sein")
    return mode


@router.post("/items/import", dependencies=[Depends(require_owner_or_admin)])
async def import_items(
    file: UploadFile = File(...),
    ctx: TenantContext = Depends(get_tenant_context),
    db: AsyncSession = Depends(get_db),
) -> dict:
    import csv
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail={"error": {"code": "invalid_encoding", "message": "CSV nicht UTF-8"}})

    reader = csv.DictReader(text.splitlines())
    if not reader.fieldnames:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "invalid_csv", "message": "CSV hat keine Header-Zeile"}},
        )

    missing_columns = [col for col in CSV_COLUMNS if col not in reader.fieldnames]
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail={"error": {"code": "missing_columns", "message": f"Fehlende Spalten: {', '.join(missing_columns)}"}},
        )

    imported = 0
    updated = 0
    errors: list[dict[str, str]] = []

    for idx, row in enumerate(reader, start=2):  # Start bei 2 wegen Header
        try:
            sku_raw = row.get("sku", "").strip()
            barcode = row.get("barcode", "").strip()
            name = row.get("name", "").strip()
            if not sku_raw or not barcode or not name:
                raise ValueError("sku, barcode und name sind Pflicht")

            normalized_sku = _normalize_sku(sku_raw)
            category_name = row.get("category", "")
            category = await _category_by_name(db=db, ctx=ctx, name=category_name)
            if category_name and category is None:
                raise ValueError(f"Kategorie '{category_name}' nicht gefunden")

            quantity = int(row.get("qty") or 0)
            unit = row.get("unit") or "pcs"
            is_active = _parse_bool(row.get("is_active", "true"))
            min_stock = int(row.get("min_stock") or 0)
            max_stock = int(row.get("max_stock") or 0)
            target_stock = int(row.get("target_stock") or 0)
            recommended_stock = int(row.get("recommended_stock") or 0)
            order_mode = _validate_order_mode(row.get("order_mode") or "0")

            item = await db.scalar(
                select(Item).where(Item.tenant_id == ctx.tenant.id, Item.sku == normalized_sku)
            )
            if item:
                if item.is_admin_created:
                    errors.append({"row": str(idx), "error": "Admin-Artikel sind schreibgeschützt"})
                    continue
                item.barcode = barcode
                item.name = name
                item.description = row.get("description") or ""
                item.category_id = category.id if category else None
                item.quantity = quantity
                item.unit = unit
                item.is_active = is_active
                item.min_stock = min_stock
                item.max_stock = max_stock
                item.target_stock = target_stock
                item.recommended_stock = recommended_stock
                item.order_mode = order_mode
                updated += 1
            else:
                new_item = Item(
                    tenant_id=ctx.tenant.id,
                    sku=normalized_sku,
                    barcode=barcode,
                    name=name,
                    description=row.get("description") or "",
                    category_id=category.id if category else None,
                    quantity=quantity,
                    unit=unit,
                    is_active=is_active,
                    min_stock=min_stock,
                    max_stock=max_stock,
                    target_stock=target_stock,
                    recommended_stock=recommended_stock,
                    order_mode=order_mode,
                    is_admin_created=False,
                )
                db.add(new_item)
                imported += 1
        except Exception as e:  # noqa: BLE001
            errors.append({"row": str(idx), "error": str(e)})

    await db.commit()
    return {"imported": imported, "updated": updated, "errors": errors}


@router.get("/items/export")
async def export_items(
    ctx: TenantContext = Depends(get_tenant_context),
    user_ctx: CurrentUserContext = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> dict:
    import csv
    from io import StringIO

    rows = (
        await db.scalars(
            select(Item)
            .where(Item.tenant_id == ctx.tenant.id)
            .order_by(Item.name.asc())
        )
    ).all()
    category_ids = {row.category_id for row in rows if row.category_id}
    categories: dict[Any, Category] = {}
    if category_ids:
        cat_rows = (await db.scalars(select(Category).where(Category.id.in_(category_ids)))).all()
        categories = {c.id: c for c in cat_rows}

    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_COLUMNS)
    writer.writeheader()
    for item in rows:
        writer.writerow(
            {
                "sku": item.sku,
                "barcode": item.barcode,
                "name": item.name,
                "description": item.description,
                "qty": item.quantity,
                "unit": item.unit,
                "is_active": item.is_active,
                "category": categories.get(item.category_id).name if item.category_id else "",
                "min_stock": item.min_stock,
                "max_stock": item.max_stock,
                "target_stock": item.target_stock,
                "recommended_stock": item.recommended_stock,
                "order_mode": item.order_mode,
            }
        )

    return {"csv": buf.getvalue()}
